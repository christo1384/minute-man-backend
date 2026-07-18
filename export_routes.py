"""
Minute Man — multi-format export utilities (FastAPI). v3: two templates.

  POST /export/pdf   -> application/pdf download
  POST /export/excel -> .xlsx download

Both accept the same structured payload (MinutesExportRequest — the output of
/api/minutes) and render it into a clean, tabular document. They do NOT call the
LLM; they only format already-extracted structured data.

v3 template behaviour:
  template="safety"  (default — a v2.2 payload with no template field lands
                      here and still works):
      PDF sections 1-6: Incidents Reviewed, Hazards & Risk Controls,
      Action Register, Decisions Made, Minutes Summary, Attendance Record.
      Excel adds an "Incidents Reviewed" sheet and puts the summary on the
      Summary cover sheet.
  template="general":
      PDF sections 1-5: Meeting Details, Attendees, Meeting Summary,
      Action Register, Decisions Made. No hazard/incident sections at all.
      Excel mirrors the same five sections.
"""

import io
from datetime import date
from typing import List, Optional

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/export", tags=["export"])

# Brand palette (matches the front-end theme)
SLATE_INDUSTRIAL = "2F4F4F"
SAFETY_ORANGE = "FF8C00"


# ---------------------------------------------------------------------------
# Shared schema — the single contract between engine, front-end, and exporter.
# v3 additions all default so a v2.2 payload validates unchanged.
# ---------------------------------------------------------------------------
class HazardRow(BaseModel):
    hazard: str
    control: str
    control_tier: str
    compliance_note: str
    custom: Optional[dict] = None  # v5: uploaded-template extra columns


class IncidentRow(BaseModel):
    description: str
    severity: str = "not stated"
    outcome: str = ""


class ActionRow(BaseModel):
    who: str
    what: str
    by_when: str
    custom: Optional[dict] = None  # v5: uploaded-template extra columns


class CarriedOverRow(BaseModel):
    """v4 — an open action from a previous meeting at this site, shown on the
    record as still outstanding (NOT a new action of this meeting)."""

    who: str = ""
    what: str
    original_date: str = ""   # the meeting_date it was raised
    by_when: str = ""


class AttendanceEntry(BaseModel):
    name: str
    signature: str = ""


class MinutesExportRequest(BaseModel):
    meeting_type: str
    site_name: Optional[str] = ""
    meeting_date: Optional[str] = Field(default_factory=lambda: date.today().isoformat())
    template: str = "safety"          # v3: "safety" | "general" (default keeps v2.2 behaviour)
    led_by: Optional[str] = ""        # v3: shown in the General "Meeting Details" section
    incidents: List[IncidentRow] = [] # v3: safety template — incidents reviewed
    summary: str = ""                 # v3: minutes summary (both templates)
    topics: List[str] = []            # v3: general template topic labels
    hazards: List[HazardRow] = []
    actions: List[ActionRow] = []
    carried_over: List[CarriedOverRow] = []  # v4: outstanding from previous meetings
    decisions: List[str] = []
    attendance: List[AttendanceEntry] = []
    template_id: Optional[int] = None        # v5: uploaded templates render spec-driven
    custom_sections: dict = {}               # v5: {"section_key": [ {row}, … ]}


DISCLAIMER = (
    "This record is AI-generated decision support based on a verbal transcript. "
    "It is not a certified legal or WorkSafe NZ compliance document. A responsible "
    "person must review, correct, and formally sign off this record before "
    "distribution or filing."
)


# ---------------------------------------------------------------------------
# PDF export (reportlab)
# ---------------------------------------------------------------------------
def _build_pdf(payload: MinutesExportRequest) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Minute Man - {payload.meeting_type}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("MMTitle", parent=styles["Title"],
                                 textColor=colors.HexColor(f"#{SLATE_INDUSTRIAL}"), fontSize=18)
    heading_style = ParagraphStyle("MMHeading", parent=styles["Heading2"],
                                   textColor=colors.HexColor(f"#{SLATE_INDUSTRIAL}"), spaceBefore=14)
    meta_style = ParagraphStyle("MMMeta", parent=styles["Normal"], textColor=colors.grey)
    body_style = styles["Normal"]
    disclaimer_style = ParagraphStyle("MMDisclaimer", parent=styles["Normal"],
                                      fontSize=8, textColor=colors.grey, spaceBefore=16)

    elements = [
        Paragraph(f"Minute Man — {payload.meeting_type} Minutes", title_style),
        Paragraph(f"Site: {payload.site_name or 'Not specified'} &nbsp;|&nbsp; Date: {payload.meeting_date}", meta_style),
        Spacer(1, 10),
    ]
    header_fill = colors.HexColor(f"#{SLATE_INDUSTRIAL}")
    accent_fill = colors.HexColor(f"#{SAFETY_ORANGE}")

    def styled_table(rows, col_widths, fill):
        # Wrap body cells in Paragraphs so long text wraps instead of clipping.
        wrapped = [rows[0]]
        for r in rows[1:]:
            wrapped.append([Paragraph(str(c), body_style) if not isinstance(c, Paragraph) else c for c in r])
        table = Table(wrapped, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), fill),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        return table

    def attendance_section(number: int, heading: str):
        elements.append(Paragraph(f"{number}. {heading}", heading_style))
        if payload.attendance:
            rows = [["Name", "Signature"]]
            rows += [[e.name, e.signature or "—"] for e in payload.attendance]
            elements.append(styled_table(rows, [80 * mm, 78 * mm], header_fill))
        else:
            elements.append(Paragraph("No attendance recorded.", body_style))

    def actions_section(number: int):
        elements.append(Paragraph(f"{number}. Action Register (Who / What / When)", heading_style))
        if payload.actions:
            rows = [["Who", "What", "By When"]]
            rows += [[a.who, a.what, a.by_when] for a in payload.actions]
            elements.append(styled_table(rows, [38 * mm, 90 * mm, 30 * mm], accent_fill))
        else:
            elements.append(Paragraph("No actions recorded.", body_style))
        # v4: carried-over actions follow the register, deliberately UNnumbered
        # so the v3 numbered-section layout (and its tests) stay intact.
        if payload.carried_over:
            elements.append(Paragraph("Outstanding Actions (carried over)", heading_style))
            rows = [["Who", "What", "Raised At", "By When"]]
            rows += [[c.who, c.what, c.original_date, c.by_when] for c in payload.carried_over]
            elements.append(styled_table(rows, [34 * mm, 72 * mm, 28 * mm, 24 * mm], accent_fill))

    def decisions_section(number: int):
        elements.append(Paragraph(f"{number}. Decisions Made", heading_style))
        if payload.decisions:
            for d in payload.decisions:
                elements.append(Paragraph(f"• {d}", body_style))
        else:
            elements.append(Paragraph("No formal decisions recorded.", body_style))

    if payload.template == "general":
        # ----- General Meeting: sections 1-5, no hazards/incidents -----
        elements.append(Paragraph("1. Meeting Details", heading_style))
        rows = [["Date", "Meeting Type", "Site / Location", "Led By"],
                [payload.meeting_date or "—", payload.meeting_type or "—",
                 payload.site_name or "—", payload.led_by or "—"]]
        elements.append(styled_table(rows, [35 * mm, 45 * mm, 45 * mm, 33 * mm], header_fill))

        attendance_section(2, "Attendees")

        elements.append(Paragraph("3. Meeting Summary", heading_style))
        if payload.summary:
            for para in payload.summary.split("\n\n"):
                elements.append(Paragraph(para.replace("\n", "<br/>"), body_style))
                elements.append(Spacer(1, 4))
        else:
            elements.append(Paragraph("No summary recorded.", body_style))

        actions_section(4)
        decisions_section(5)
    else:
        # ----- Safety / Toolbox Talk: sections 1-6 -----
        elements.append(Paragraph("1. Incidents Reviewed", heading_style))
        if payload.incidents:
            rows = [["Description", "Severity", "Review Outcome"]]
            rows += [[i.description, i.severity, i.outcome] for i in payload.incidents]
            elements.append(styled_table(rows, [70 * mm, 28 * mm, 60 * mm], header_fill))
        else:
            elements.append(Paragraph("No incidents reviewed.", body_style))

        elements.append(Paragraph("2. Hazards &amp; Risk Controls", heading_style))
        if payload.hazards:
            rows = [["Hazard Identified", "Control Discussed", "Hierarchy Tier", "HSWA Compliance Note"]]
            rows += [[h.hazard, h.control, h.control_tier, h.compliance_note] for h in payload.hazards]
            elements.append(styled_table(rows, [42 * mm, 42 * mm, 32 * mm, 42 * mm], header_fill))
        else:
            elements.append(Paragraph("No hazards recorded.", body_style))

        actions_section(3)
        decisions_section(4)

        elements.append(Paragraph("5. Minutes Summary", heading_style))
        if payload.summary:
            for para in payload.summary.split("\n\n"):
                elements.append(Paragraph(para.replace("\n", "<br/>"), body_style))
                elements.append(Spacer(1, 4))
        else:
            elements.append(Paragraph("No summary recorded.", body_style))

        attendance_section(6, "Attendance Record")

    elements.append(Paragraph(DISCLAIMER, disclaimer_style))
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# v5 — spec-driven rendering for UPLOADED templates: sheet order, sheet titles
# and column labels come from the template's spec, so the export looks like
# the spreadsheet the user uploaded. Builtin templates keep their dedicated
# v4 renderers above (regression-tested byte-comparable) — the generic path
# is only taken when payload.template_id points at an uploaded template.
# ---------------------------------------------------------------------------
def _uploaded_spec_for(payload: MinutesExportRequest) -> dict | None:
    if payload.template_id is None:
        return None
    import crud
    from db import SessionLocal

    with SessionLocal() as session:
        t = crud.get_template(session, payload.template_id)
    if t is None or (t.extra or {}).get("builtin_key"):
        return None  # unknown or builtin → dedicated paths
    return t.spec or {}


def _spec_rows(payload: MinutesExportRequest, section: dict):
    """(header labels, data rows) for one spec section, from the payload."""
    kind = section.get("kind")
    cols = section.get("columns", [])
    labels = [c["label"] for c in cols]

    def cell(row_obj, col):
        if col.get("maps_to"):
            return getattr(row_obj, col["maps_to"], "") or ""
        return (row_obj.custom or {}).get(col.get("key"), "") if getattr(row_obj, "custom", None) else ""

    if kind == "hazards":
        return labels, [[cell(h, c) for c in cols] for h in payload.hazards]
    if kind == "actions":
        return labels, [[cell(a, c) for c in cols] for a in payload.actions]
    if kind == "decisions":
        return labels, [[d] for d in payload.decisions]
    if kind == "attendance":
        def att_cell(e, c):
            key = c.get("maps_to")
            return getattr(e, key, "") or "" if key else ""
        return labels, [[att_cell(e, c) for c in cols] for e in payload.attendance]
    if kind == "custom":
        from template_engine import _snake
        rows = (payload.custom_sections or {}).get(_snake(section.get("title", "")), [])
        return labels, [[str(r.get(c["key"], "") or "") for c in cols] for r in rows]
    return labels, []


def _summary_pairs(payload: MinutesExportRequest, section: dict):
    known = {"meeting_type": payload.meeting_type, "site_name": payload.site_name,
             "meeting_date": payload.meeting_date, "led_by": payload.led_by}
    pairs = []
    for f in section.get("fields", []):
        val = known.get(f["key"])
        if val is None:
            val = (payload.custom_sections or {}).get("_summary", {}).get(f["key"], "")
        pairs.append((f["label"], val or ""))
    return pairs


def _build_pdf_from_spec(payload: MinutesExportRequest, spec: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Minute Man - {payload.meeting_type}")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("MMTitle", parent=styles["Title"],
                                 textColor=colors.HexColor(f"#{SLATE_INDUSTRIAL}"), fontSize=18)
    heading_style = ParagraphStyle("MMHeading", parent=styles["Heading2"],
                                   textColor=colors.HexColor(f"#{SLATE_INDUSTRIAL}"), spaceBefore=14)
    meta_style = ParagraphStyle("MMMeta", parent=styles["Normal"], textColor=colors.grey)
    body_style = styles["Normal"]
    disclaimer_style = ParagraphStyle("MMDisclaimer", parent=styles["Normal"],
                                      fontSize=8, textColor=colors.grey, spaceBefore=16)
    header_fill = colors.HexColor(f"#{SLATE_INDUSTRIAL}")
    accent_fill = colors.HexColor(f"#{SAFETY_ORANGE}")

    elements = [
        Paragraph(f"Minute Man — {payload.meeting_type} Minutes", title_style),
        Paragraph(f"Site: {payload.site_name or 'Not specified'} &nbsp;|&nbsp; Date: {payload.meeting_date}", meta_style),
        Spacer(1, 10),
    ]

    def styled_table(rows, fill):
        wrapped = [rows[0]]
        for r in rows[1:]:
            wrapped.append([Paragraph(str(c), body_style) for c in r])
        n = max(len(rows[0]), 1)
        table = Table(wrapped, colWidths=[(158 * mm) / n] * n, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), fill),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        return table

    num = 0
    for section in spec.get("sections", []):
        num += 1
        title = section.get("title", "")
        elements.append(Paragraph(f"{num}. {title}".replace("&", "&amp;"), heading_style))
        if section.get("kind") == "summary":
            rows = [["Field", "Value"]] + [[a, b] for a, b in _summary_pairs(payload, section)]
            elements.append(styled_table(rows, header_fill))
            if payload.summary:
                elements.append(Spacer(1, 4))
                elements.append(Paragraph(payload.summary.replace("\n", "<br/>"), body_style))
            continue
        labels, data = _spec_rows(payload, section)
        if data:
            fill = accent_fill if section.get("kind") == "actions" else header_fill
            elements.append(styled_table([labels] + data, fill))
        else:
            elements.append(Paragraph("Nothing recorded.", body_style))
        if section.get("kind") == "actions" and payload.carried_over:
            elements.append(Paragraph("Outstanding Actions (carried over)", heading_style))
            rows = [["Who", "What", "Raised At", "By When"]]
            rows += [[c.who, c.what, c.original_date, c.by_when] for c in payload.carried_over]
            elements.append(styled_table(rows, accent_fill))

    elements.append(Paragraph(DISCLAIMER, disclaimer_style))
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _build_excel_from_spec(payload: MinutesExportRequest, spec: dict) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    fill_slate = PatternFill(start_color=SLATE_INDUSTRIAL, end_color=SLATE_INDUSTRIAL, fill_type="solid")
    fill_orange = PatternFill(start_color=SAFETY_ORANGE, end_color=SAFETY_ORANGE, fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    first = True

    def sheet(title):
        nonlocal first
        safe = title[:31].replace("/", "-").replace("\\", "-").replace("*", "-") \
                         .replace("?", "-").replace("[", "(").replace("]", ")").replace(":", "-")
        if first:
            ws = wb.active
            ws.title = safe
            first = False
            return ws
        return wb.create_sheet(safe)

    for section in spec.get("sections", []):
        ws = sheet(section.get("title", "Sheet"))
        if section.get("kind") == "summary":
            ws.append(["Minute Man — Meeting Minutes"])
            ws["A1"].font = Font(bold=True, size=14, color=SLATE_INDUSTRIAL)
            for a, b in _summary_pairs(payload, section):
                ws.append([a, b])
            if payload.summary:
                ws.append([])
                ws.append([payload.summary])
                ws.cell(row=ws.max_row, column=1).alignment = wrap
            ws.append([])
            ws.append([DISCLAIMER])
            ws.cell(row=ws.max_row, column=1).alignment = wrap
            _autosize(ws, [40, 50])
            continue
        labels, data = _spec_rows(payload, section)
        ws.append(labels)
        fill = fill_orange if section.get("kind") == "actions" else fill_slate
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
        for row in data:
            ws.append(row)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        _autosize(ws, [max(14, min(50, 160 // max(len(labels), 1)))] * max(len(labels), 1))
        if section.get("kind") == "actions" and payload.carried_over:
            wsc = sheet("Outstanding (carried over)")
            wsc.append(["Who", "What", "Raised At", "By When"])
            for cell in wsc[1]:
                cell.font = header_font
                cell.fill = fill_orange
            for c2 in payload.carried_over:
                wsc.append([c2.who, c2.what, c2.original_date, c2.by_when])
            _autosize(wsc, [24, 50, 18, 18])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@router.post("/pdf")
def export_pdf(payload: MinutesExportRequest):
    spec = _uploaded_spec_for(payload)  # v5: uploaded templates render spec-driven
    pdf_bytes = _build_pdf_from_spec(payload, spec) if spec else _build_pdf(payload)
    filename = f"minute-man-{payload.meeting_type.lower().replace(' ', '-')}-{payload.meeting_date}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Excel export (openpyxl)
# ---------------------------------------------------------------------------
def _autosize(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _build_excel(payload: MinutesExportRequest) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    fill_slate = PatternFill(start_color=SLATE_INDUSTRIAL, end_color=SLATE_INDUSTRIAL, fill_type="solid")
    fill_orange = PatternFill(start_color=SAFETY_ORANGE, end_color=SAFETY_ORANGE, fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    is_general = payload.template == "general"

    def style_header(ws, fill):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

    def wrap_body(ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

    ws_first = wb.active  # openpyxl always starts with one sheet — reuse it below

    if not is_general:
        # Sheet: Incidents Reviewed (NEW in v3, safety only)
        ws_inc = ws_first
        ws_inc.title = "Incidents Reviewed"
        ws_inc.append(["Description", "Severity", "Review Outcome"])
        style_header(ws_inc, fill_slate)
        if payload.incidents:
            for i in payload.incidents:
                ws_inc.append([i.description, i.severity, i.outcome])
        else:
            ws_inc.append(["No incidents reviewed.", "", ""])
        wrap_body(ws_inc)
        _autosize(ws_inc, [50, 18, 50])

        # Sheet: Hazards & Controls
        ws1 = wb.create_sheet("Hazards & Controls")
        ws1.append(["Hazard Identified", "Control Discussed", "Hierarchy of Controls Tier", "HSWA Compliance Note"])
        style_header(ws1, fill_slate)
        for h in payload.hazards:
            ws1.append([h.hazard, h.control, h.control_tier, h.compliance_note])
        wrap_body(ws1)
        _autosize(ws1, [32, 32, 26, 34])
    else:
        # General template reuses the first sheet as the Action Register.
        ws_first.title = "Action Register"

    # Sheet: Action Register (both templates)
    ws2 = ws_first if is_general else wb.create_sheet("Action Register")
    ws2.append(["Who", "What", "By When"])
    style_header(ws2, fill_orange)
    for a in payload.actions:
        ws2.append([a.who, a.what, a.by_when])
    wrap_body(ws2)
    _autosize(ws2, [24, 60, 18])

    # v4: Outstanding (carried over) sheet — only when there is content
    if payload.carried_over:
        wsc = wb.create_sheet("Outstanding (carried over)")
        wsc.append(["Who", "What", "Raised At", "By When"])
        style_header(wsc, fill_orange)
        for c2 in payload.carried_over:
            wsc.append([c2.who, c2.what, c2.original_date, c2.by_when])
        wrap_body(wsc)
        _autosize(wsc, [24, 50, 18, 18])

    # Sheet: Decisions (both templates)
    ws3 = wb.create_sheet("Decisions")
    ws3.append(["Decision"])
    style_header(ws3, fill_slate)
    if payload.decisions:
        for d in payload.decisions:
            ws3.append([d])
    else:
        ws3.append(["No formal decisions recorded."])
    _autosize(ws3, [80])

    # Sheet: Attendance (both templates; "Attendees" label for general)
    ws4 = wb.create_sheet("Attendees" if is_general else "Attendance Record")
    ws4.append(["Name", "Signature"])
    style_header(ws4, fill_slate)
    for e in payload.attendance:
        ws4.append([e.name, e.signature])
    _autosize(ws4, [30, 30])

    # Sheet 0: Summary cover (meeting details + v3 minutes summary)
    ws0 = wb.create_sheet("Summary", 0)
    ws0.append(["Minute Man — Meeting Minutes"])
    ws0["A1"].font = Font(bold=True, size=14, color=SLATE_INDUSTRIAL)
    ws0.append(["Meeting Type", payload.meeting_type])
    ws0.append(["Site", payload.site_name or "Not specified"])
    ws0.append(["Date", payload.meeting_date])
    if payload.led_by:
        ws0.append(["Led By", payload.led_by])
    if is_general and payload.topics:
        ws0.append(["Topics", ", ".join(payload.topics)])
    ws0.append([])
    if payload.summary:
        ws0.append(["Minutes Summary"])
        ws0[f"A{ws0.max_row}"].font = Font(bold=True, color=SLATE_INDUSTRIAL)
        ws0.append([payload.summary])
        ws0[f"A{ws0.max_row}"].alignment = wrap
        ws0.append([])
    ws0.append([DISCLAIMER])
    ws0[f"A{ws0.max_row}"].alignment = wrap
    _autosize(ws0, [90])
    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@router.post("/excel")
def export_excel(payload: MinutesExportRequest):
    spec = _uploaded_spec_for(payload)  # v5: uploaded templates render spec-driven
    xlsx_bytes = _build_excel_from_spec(payload, spec) if spec else _build_excel(payload)
    filename = f"minute-man-{payload.meeting_type.lower().replace(' ', '-')}-{payload.meeting_date}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# v4 — Action Register export: the current filtered list as one styled sheet.
# Same filters as GET /api/actions, passed in the POST body.
# ---------------------------------------------------------------------------
class ActionsExportRequest(BaseModel):
    status: str = "open"                 # "open" | "closed" | "all"
    who: Optional[str] = None
    site_name: Optional[str] = None
    template: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    overdue: bool = False
    limit: int = 500
    offset: int = 0


@router.post("/actions.xlsx")
def export_actions(req: ActionsExportRequest):
    import crud
    from db import SessionLocal

    with SessionLocal() as session:
        rows, _total = crud.list_actions(
            session, status=req.status if req.status in ("open", "closed", "all") else "open",
            who=req.who, site_name=req.site_name, template=req.template,
            date_from=req.date_from, date_to=req.date_to, overdue=req.overdue,
            limit=min(req.limit, 500), offset=req.offset)

    wb = Workbook()
    ws = wb.active
    ws.title = "Action Register"
    header_font = Font(bold=True, color="FFFFFF")
    fill_orange = PatternFill(start_color=SAFETY_ORANGE, end_color=SAFETY_ORANGE, fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    red_font = Font(color="C02626", bold=True)

    ws.append(["Status", "Overdue", "Who", "What", "By When", "Due Date",
               "Site", "Meeting", "Meeting Date", "Closed By", "Closed At"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill_orange
    for r in rows:
        ws.append([r["status"], "YES" if r["overdue"] else "", r["who"], r["what"],
                   r["by_when"], r["due_date"] or "", r["site_name"],
                   r["meeting_type"], r["meeting_date"], r["closed_by"] or "",
                   (r["closed_at"] or "")[:19].replace("T", " ")])
        if r["overdue"]:
            ws.cell(row=ws.max_row, column=2).font = red_font
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    _autosize(ws, [9, 9, 18, 46, 14, 12, 22, 18, 14, 14, 18])
    ws.append([])
    ws.append([DISCLAIMER])
    ws.cell(row=ws.max_row, column=1).alignment = wrap

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"minute-man-action-register-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
